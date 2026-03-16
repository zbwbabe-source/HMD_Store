from __future__ import annotations

import argparse
import csv
import re
import sqlite3
import subprocess
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DIMENSION_COLUMNS = [
    "brand",
    "country",
    "channel",
    "store_code",
    "store_name",
    "account_name",
    "account_name_clean",
    "account_level",
    "parent_account_name",
    "account_group_name",
]

ANNUAL_TOTAL_SUFFIX = "년 합계"
SPECIAL_CHILD_PARENT_MAP = {
    "매출원가": "매출원가합계",
    "운반비": "매출원가합계",
}


@dataclass(frozen=True)
class PeriodColumn:
    index: int
    header: str
    period_type: str
    year: int
    month: int | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize store P&L Excel data into CSV files and a SQLite database."
    )
    parser.add_argument("--input", default="Store_Rawdata.xlsx", help="Path to the source xlsx file.")
    parser.add_argument("--sheet", default="Sheet1", help="Worksheet name to import.")
    parser.add_argument("--output-dir", default="data/normalized", help="Directory to write normalized CSV files.")
    parser.add_argument("--db-path", default="data/store_dashboard.sqlite", help="Path to the SQLite database file.")
    return parser.parse_args()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_decimal(value: object) -> str:
    if value is None:
        return "0"
    text = str(value).strip().replace(",", "")
    if text == "" or text.startswith("#"):
        return "0"
    try:
        return str(Decimal(text))
    except InvalidOperation as exc:
        raise ValueError(f"Unable to parse numeric value: {value!r}") from exc


def detect_period_columns(headers: list[object]) -> list[PeriodColumn]:
    period_columns: list[PeriodColumn] = []
    monthly_pattern = re.compile(r"^(\d{2})(\d{2})$")

    for index, raw_header in enumerate(headers[6:], start=7):
        header = normalize_text(raw_header)
        monthly_match = monthly_pattern.match(header)
        if monthly_match:
            year = 2000 + int(monthly_match.group(1))
            month = int(monthly_match.group(2))
            period_columns.append(
                PeriodColumn(index=index, header=header, period_type="monthly", year=year, month=month)
            )
            continue

        if header.endswith(ANNUAL_TOTAL_SUFFIX) and len(header) >= 2 and header[:2].isdigit():
            year = 2000 + int(header[:2])
            period_columns.append(
                PeriodColumn(index=index, header=header, period_type="annual_total", year=year, month=None)
            )

    return period_columns


def normalize_account_name(account_name: str) -> str:
    return re.sub(r"^[-\s]+", "", account_name).strip()


def derive_account_hierarchy(account_name: str, current_parent: str) -> tuple[str, str, str, str]:
    account_name_clean = normalize_account_name(account_name)

    if account_name.startswith("-") or account_name.startswith(" -"):
        parent_account_name = current_parent
        account_level = "child"
        account_group_name = parent_account_name or account_name_clean
        return account_name_clean, account_level, parent_account_name, account_group_name

    if account_name in SPECIAL_CHILD_PARENT_MAP:
        parent_account_name = SPECIAL_CHILD_PARENT_MAP[account_name]
        account_level = "child"
        account_group_name = parent_account_name
        return account_name_clean, account_level, parent_account_name, account_group_name

    account_level = "parent"
    parent_account_name = ""
    account_group_name = account_name_clean
    return account_name_clean, account_level, parent_account_name, account_group_name


def iter_records(ws) -> tuple[list[dict[str, str]], list[dict[str, str]], int]:
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    period_columns = detect_period_columns(headers)

    monthly_records: list[dict[str, str]] = []
    annual_records: list[dict[str, str]] = []
    skipped_rows = 0
    current_store_key: tuple[str, str, str, str, str] | None = None
    current_parent_account = ""

    for excel_row_number, row in enumerate(rows, start=2):
        store_dims = (
            normalize_text(row[0]),
            normalize_text(row[1]),
            normalize_text(row[2]),
            normalize_text(row[3]),
            normalize_text(row[4]),
        )
        account_name = normalize_text(row[5])

        if not any(store_dims) and account_name == "":
            skipped_rows += 1
            continue

        if normalize_text(row[0]) == normalize_text(headers[0]) or account_name == normalize_text(headers[5]):
            skipped_rows += 1
            current_parent_account = ""
            current_store_key = None
            continue

        if current_store_key != store_dims:
            current_store_key = store_dims
            current_parent_account = ""

        account_name_clean, account_level, parent_account_name, account_group_name = derive_account_hierarchy(
            account_name,
            current_parent_account,
        )

        dims = {
            "brand": store_dims[0],
            "country": store_dims[1],
            "channel": store_dims[2],
            "store_code": store_dims[3],
            "store_name": store_dims[4],
            "account_name": account_name,
            "account_name_clean": account_name_clean,
            "account_level": account_level,
            "parent_account_name": parent_account_name,
            "account_group_name": account_group_name,
        }

        if account_level == "parent":
            current_parent_account = account_name_clean

        for period in period_columns:
            amount = parse_decimal(row[period.index - 1])
            if period.period_type == "monthly":
                monthly_records.append(
                    {
                        **dims,
                        "year": str(period.year),
                        "month": str(period.month),
                        "period_key": f"{period.year:04d}-{period.month:02d}",
                        "period_date": date(period.year, period.month, 1).isoformat(),
                        "amount": amount,
                        "source_column": period.header,
                        "source_row": str(excel_row_number),
                    }
                )
            else:
                annual_records.append(
                    {
                        **dims,
                        "year": str(period.year),
                        "amount": amount,
                        "source_column": period.header,
                        "source_row": str(excel_row_number),
                    }
                )

    return monthly_records, annual_records, skipped_rows


def write_csv(path: Path, fieldnames: Iterable[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        writer.writerows(rows)


def reset_database(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        DROP TABLE IF EXISTS monthly_pnl;
        DROP TABLE IF EXISTS annual_pnl;
        DROP TABLE IF EXISTS stores;
        DROP TABLE IF EXISTS accounts;
        DROP TABLE IF EXISTS import_batches;

        CREATE TABLE import_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            imported_at TEXT NOT NULL,
            monthly_row_count INTEGER NOT NULL,
            annual_row_count INTEGER NOT NULL,
            skipped_row_count INTEGER NOT NULL
        );

        CREATE TABLE stores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            brand TEXT NOT NULL,
            country TEXT NOT NULL,
            channel TEXT NOT NULL,
            store_code TEXT NOT NULL UNIQUE,
            store_name TEXT NOT NULL,
            open_date TEXT,
            close_date TEXT,
            status TEXT NOT NULL DEFAULT 'active'
        );

        CREATE TABLE accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account_name TEXT NOT NULL,
            account_name_clean TEXT NOT NULL,
            account_level TEXT NOT NULL,
            parent_account_name TEXT NOT NULL,
            account_group_name TEXT NOT NULL,
            UNIQUE (
                account_name,
                account_name_clean,
                account_level,
                parent_account_name,
                account_group_name
            )
        );

        CREATE TABLE monthly_pnl (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_batch_id INTEGER NOT NULL,
            store_id INTEGER NOT NULL,
            account_id INTEGER NOT NULL,
            brand TEXT NOT NULL,
            country TEXT NOT NULL,
            channel TEXT NOT NULL,
            store_code TEXT NOT NULL,
            store_name TEXT NOT NULL,
            account_name TEXT NOT NULL,
            account_name_clean TEXT NOT NULL,
            account_level TEXT NOT NULL,
            parent_account_name TEXT NOT NULL,
            account_group_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            period_key TEXT NOT NULL,
            period_date TEXT NOT NULL,
            amount NUMERIC NOT NULL,
            source_column TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            FOREIGN KEY (import_batch_id) REFERENCES import_batches(id),
            FOREIGN KEY (store_id) REFERENCES stores(id),
            FOREIGN KEY (account_id) REFERENCES accounts(id)
        );

        CREATE TABLE annual_pnl (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_batch_id INTEGER NOT NULL,
            store_id INTEGER NOT NULL,
            account_id INTEGER NOT NULL,
            brand TEXT NOT NULL,
            country TEXT NOT NULL,
            channel TEXT NOT NULL,
            store_code TEXT NOT NULL,
            store_name TEXT NOT NULL,
            account_name TEXT NOT NULL,
            account_name_clean TEXT NOT NULL,
            account_level TEXT NOT NULL,
            parent_account_name TEXT NOT NULL,
            account_group_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            amount NUMERIC NOT NULL,
            source_column TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            FOREIGN KEY (import_batch_id) REFERENCES import_batches(id),
            FOREIGN KEY (store_id) REFERENCES stores(id),
            FOREIGN KEY (account_id) REFERENCES accounts(id)
        );

        CREATE INDEX idx_monthly_store_period ON monthly_pnl (store_code, year, month);
        CREATE INDEX idx_monthly_account_group ON monthly_pnl (account_group_name);
        CREATE INDEX idx_monthly_store_id_period ON monthly_pnl (store_id, period_date);
        CREATE INDEX idx_monthly_account_id_period ON monthly_pnl (account_id, period_date);
        CREATE INDEX idx_monthly_import_batch ON monthly_pnl (import_batch_id);
        CREATE INDEX idx_annual_store_year ON annual_pnl (store_code, year);
        CREATE INDEX idx_annual_store_id_year ON annual_pnl (store_id, year);
        CREATE INDEX idx_annual_account_id_year ON annual_pnl (account_id, year);
        CREATE INDEX idx_annual_import_batch ON annual_pnl (import_batch_id);
        CREATE INDEX idx_stores_lookup ON stores (brand, country, channel);
        CREATE INDEX idx_accounts_group ON accounts (account_group_name, account_level);
        """
    )
    connection.commit()


def insert_rows(connection: sqlite3.Connection, table_name: str, rows: list[dict[str, str | int]]) -> None:
    if not rows:
        return

    columns = list(rows[0].keys())
    placeholders = ", ".join("?" for _ in columns)
    sql = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})"
    values = [tuple(row[column] for column in columns) for row in rows]
    connection.executemany(sql, values)
    connection.commit()


def insert_import_batch(
    connection: sqlite3.Connection,
    input_path: Path,
    sheet_name: str,
    monthly_count: int,
    annual_count: int,
    skipped_rows: int,
) -> int:
    cursor = connection.execute(
        """
        INSERT INTO import_batches (
            source_file,
            source_sheet,
            imported_at,
            monthly_row_count,
            annual_row_count,
            skipped_row_count
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            input_path.name,
            sheet_name,
            datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            monthly_count,
            annual_count,
            skipped_rows,
        ),
    )
    connection.commit()
    return int(cursor.lastrowid)


def build_store_map(connection: sqlite3.Connection, monthly_rows: list[dict[str, str]], annual_rows: list[dict[str, str]]) -> dict[str, int]:
    store_records: dict[str, tuple[str, str, str, str, str]] = {}
    for row in [*monthly_rows, *annual_rows]:
        store_code = row["store_code"]
        if store_code not in store_records:
            store_records[store_code] = (
                row["brand"],
                row["country"],
                row["channel"],
                row["store_code"],
                row["store_name"],
            )

    connection.executemany(
        """
        INSERT INTO stores (brand, country, channel, store_code, store_name)
        VALUES (?, ?, ?, ?, ?)
        """,
        list(store_records.values()),
    )
    connection.commit()

    return {
        store_code: store_id
        for store_id, store_code in connection.execute("SELECT id, store_code FROM stores")
    }


def build_account_map(
    connection: sqlite3.Connection,
    monthly_rows: list[dict[str, str]],
    annual_rows: list[dict[str, str]],
) -> dict[tuple[str, str, str, str, str], int]:
    account_records: dict[tuple[str, str, str, str, str], tuple[str, str, str, str, str]] = {}
    for row in [*monthly_rows, *annual_rows]:
        key = (
            row["account_name"],
            row["account_name_clean"],
            row["account_level"],
            row["parent_account_name"],
            row["account_group_name"],
        )
        account_records[key] = key

    connection.executemany(
        """
        INSERT INTO accounts (
            account_name,
            account_name_clean,
            account_level,
            parent_account_name,
            account_group_name
        ) VALUES (?, ?, ?, ?, ?)
        """,
        list(account_records.values()),
    )
    connection.commit()

    return {
        (
            account_name,
            account_name_clean,
            account_level,
            parent_account_name,
            account_group_name,
        ): account_id
        for account_id, account_name, account_name_clean, account_level, parent_account_name, account_group_name in connection.execute(
            """
            SELECT
                id,
                account_name,
                account_name_clean,
                account_level,
                parent_account_name,
                account_group_name
            FROM accounts
            """
        )
    }


def attach_dimension_ids(
    rows: list[dict[str, str]],
    import_batch_id: int,
    store_map: dict[str, int],
    account_map: dict[tuple[str, str, str, str, str], int],
) -> list[dict[str, str | int]]:
    enriched_rows: list[dict[str, str | int]] = []
    for row in rows:
        account_key = (
            row["account_name"],
            row["account_name_clean"],
            row["account_level"],
            row["parent_account_name"],
            row["account_group_name"],
        )
        enriched_rows.append(
            {
                "import_batch_id": import_batch_id,
                "store_id": store_map[row["store_code"]],
                "account_id": account_map[account_key],
                **row,
            }
        )
    return enriched_rows


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).resolve()
    output_dir = Path(args.output_dir).resolve()
    db_path = Path(args.db_path).resolve()

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[args.sheet]

    monthly_records, annual_records, skipped_rows = iter_records(ws)

    write_csv(
        output_dir / "monthly_pnl.csv",
        [
            *DIMENSION_COLUMNS,
            "year",
            "month",
            "period_key",
            "period_date",
            "amount",
            "source_column",
            "source_row",
        ],
        monthly_records,
    )
    write_csv(
        output_dir / "annual_pnl.csv",
        [*DIMENSION_COLUMNS, "year", "amount", "source_column", "source_row"],
        annual_records,
    )

    db_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(db_path)
    try:
        reset_database(connection)
        import_batch_id = insert_import_batch(
            connection,
            input_path,
            args.sheet,
            len(monthly_records),
            len(annual_records),
            skipped_rows,
        )
        store_map = build_store_map(connection, monthly_records, annual_records)
        account_map = build_account_map(connection, monthly_records, annual_records)
        insert_rows(
            connection,
            "monthly_pnl",
            attach_dimension_ids(monthly_records, import_batch_id, store_map, account_map),
        )
        insert_rows(
            connection,
            "annual_pnl",
            attach_dimension_ids(annual_records, import_batch_id, store_map, account_map),
        )
    finally:
        connection.close()

    export_script = Path(__file__).with_name("export_dashboard_data.py")
    if export_script.exists():
        subprocess.run([sys.executable, str(export_script)], check=True)

    print(f"Imported workbook: {input_path}")
    print(f"Monthly rows: {len(monthly_records)}")
    print(f"Annual rows: {len(annual_records)}")
    print(f"Skipped rows: {skipped_rows}")
    print(f"CSV output: {output_dir}")
    print(f"SQLite DB: {db_path}")
    if export_script.exists():
        print(f"Dashboard data: {(Path('dashboard/data.js')).resolve()}")


if __name__ == "__main__":
    main()
